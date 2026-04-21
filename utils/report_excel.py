"""
ExcelReporter — generates an Excel test report from pytest results.

Called by conftest.py hook: pytest_sessionfinish
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from utils.logger import get_logger

logger = get_logger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed — Excel report disabled")


# ── Colour palette ────────────────────────────────────────────────────────────

_GREEN  = "FF92D050"
_RED    = "FFFF0000"
_YELLOW = "FFFFC000"
_BLUE   = "FF4472C4"
_GREY   = "FFD9D9D9"
_WHITE  = "FFFFFFFF"
_PURPLE = "FFB8CCE4"  # manual tests


def _fill(hex_color: str) -> "PatternFill":
    """
    @brief  Tạo PatternFill màu solid từ mã hex
    @param  hex_color: Mã màu hex 8 ký tự ARGB (ví dụ: "FF92D050")
    @retval PatternFill — openpyxl fill object để tô màu cell
    """
    return PatternFill("solid", fgColor=hex_color)


def _bold(size: int = 11) -> "Font":
    """
    @brief  Tạo Font bold với kích thước chỉ định
    @param  size: Kích thước font (default: 11)
    @retval Font — openpyxl font object
    """
    return Font(bold=True, size=size)


def _center() -> "Alignment":
    """
    @brief  Tạo Alignment căn giữa theo cả chiều ngang và dọc, bật wrap text
    @retval Alignment — openpyxl alignment object
    """
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border() -> "Border":
    """
    @brief  Tạo Border với đường viền thin ở 4 phía của cell
    @retval Border — openpyxl border object
    """
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


# ── Data container ────────────────────────────────────────────────────────────

class TestResult:
    __slots__ = (
        "test_id", "brief", "test_level", "test_type",
        "execution_type", "hw_depend", "outcome",
        "duration", "error_message", "nodeid", "screenshot_path",
    )

    def __init__(self, **kwargs):
        """
        @brief  Khởi tạo TestResult từ keyword arguments; các field không truyền mặc định là chuỗi rỗng
        @param  kwargs: Các field của TestResult (test_id, brief, test_level, test_type, execution_type, hw_depend, outcome, duration, error_message, nodeid)
        @retval None
        """
        for slot in self.__slots__:
            setattr(self, slot, kwargs.get(slot, ""))


# ── Reporter ──────────────────────────────────────────────────────────────────

class ExcelReporter:
    """Builds an Excel workbook from a list of TestResult objects."""

    COLUMNS = [
        ("Test ID",        20),
        ("Brief",          40),
        ("Level",          12),
        ("Type",           14),
        ("Execution",      16),
        ("HW Depend",      11),
        ("Result",         10),
        ("Duration (s)",   13),
        ("Error / Notes",  50),
        ("Screenshot",     30),
    ]

    _SCREENSHOT_COL = 10   # cột J (1-based)
    _ROW_HEIGHT_PX  = 80   # chiều cao row khi có ảnh

    def __init__(self, output_path: Optional[str] = None):
        """
        @brief  Khởi tạo ExcelReporter với đường dẫn file đầu ra
        @param  output_path: Đường dẫn file Excel đầu ra; mặc định lấy từ settings.report nếu None
        @retval None
        """
        from config import get_settings
        cfg = get_settings().report
        if output_path:
            self.output_path = Path(output_path)
        else:
            self.output_path = Path(cfg.excel_dir) / cfg.excel_filename

    def generate(self, results: List[TestResult]) -> Path:
        """
        @brief  Tạo file Excel report với 2 sheets: Summary và Test Results
        @param  results: Danh sách TestResult từ session pytest
        @retval Path — đường dẫn file Excel đã tạo
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("Cannot generate Excel report: openpyxl not installed")
            return self.output_path

        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        self._build_summary_sheet(wb, results)
        self._build_details_sheet(wb, results)

        wb.save(self.output_path)
        logger.info(f"Excel report saved: {self.output_path}")
        return self.output_path

    # ── Summary sheet ─────────────────────────────────────────────────────────

    def _build_summary_sheet(self, wb: "Workbook", results: List[TestResult]) -> None:
        """
        @brief  Tạo sheet "Summary" với thống kê tổng hợp: Total, PASSED, FAILED, SKIPPED, ERROR, pass rate
        @param  wb: Workbook openpyxl đang được build
        @param  results: Danh sách TestResult cần thống kê
        @retval None
        """
        ws = wb.active
        ws.title = "Summary"

        total   = len(results)
        passed  = sum(1 for r in results if r.outcome == "passed")
        failed  = sum(1 for r in results if r.outcome == "failed")
        skipped = sum(1 for r in results if r.outcome == "skipped")
        errors  = sum(1 for r in results if r.outcome == "error")
        manual  = sum(1 for r in results if r.outcome == "manual")
        auto_total = total - manual  # chỉ tính pass rate trên các TC automatic

        rows = [
            ("Report generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total",   total),
            ("PASSED",  passed),
            ("FAILED",  failed),
            ("SKIPPED", skipped),
            ("ERROR",   errors),
            ("MANUAL",  manual),
            ("Pass rate", f"{(passed/auto_total*100):.1f}%" if auto_total else "N/A"),
        ]

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20

        for r_idx, (label, value) in enumerate(rows, start=1):
            label_cell = ws.cell(r_idx, 1, label)
            value_cell = ws.cell(r_idx, 2, value)
            label_cell.font = _bold()
            label_cell.fill = _fill(_GREY)
            label_cell.border = _thin_border()
            label_cell.alignment = _center()
            value_cell.border = _thin_border()
            value_cell.alignment = _center()

            if label == "PASSED":
                value_cell.fill = _fill(_GREEN)
            elif label in ("FAILED", "ERROR"):
                value_cell.fill = _fill(_RED)
            elif label == "SKIPPED":
                value_cell.fill = _fill(_YELLOW)
            elif label == "MANUAL":
                value_cell.fill = _fill(_PURPLE)

    # ── Details sheet ─────────────────────────────────────────────────────────

    def _build_details_sheet(self, wb: "Workbook", results: List[TestResult]) -> None:
        """
        @brief  Tạo sheet "Test Results" với header màu xanh và mỗi row tô màu theo outcome
        @param  wb: Workbook openpyxl đang được build
        @param  results: Danh sách TestResult cần ghi vào sheet
        @retval None
        """
        ws = wb.create_sheet("Test Results")

        # Header row
        for col_idx, (header, width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFFFF", size=11)
            cell.fill = _fill(_BLUE)
            cell.alignment = _center()
            cell.border = _thin_border()
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 24

        # Data rows
        for r_idx, result in enumerate(results, start=2):
            outcome = result.outcome.upper()
            row_fill = {
                "PASSED":  _fill(_GREEN),
                "FAILED":  _fill(_RED),
                "SKIPPED": _fill(_YELLOW),
                "MANUAL":  _fill(_PURPLE),
            }.get(outcome, _fill(_WHITE))

            values = [
                result.test_id,
                result.brief,
                result.test_level,
                result.test_type,
                result.execution_type,
                "Yes" if result.hw_depend else "No",
                outcome,
                result.duration if result.duration == "—" else (f"{float(result.duration):.2f}" if result.duration else ""),
                result.error_message,
                "",  # placeholder cho cột Screenshot
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(r_idx, col_idx, value)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col_idx == 7:  # Result column → coloured
                    cell.fill = row_fill
                    cell.font = _bold()
                    cell.alignment = _center()

            # Nhúng ảnh thumbnail vào cột Screenshot
            screenshot = getattr(result, "screenshot_path", "")
            if screenshot and Path(screenshot).exists():
                try:
                    from openpyxl.drawing.image import Image as XLImage
                    from PIL import Image as PILImage

                    # Resize về thumbnail trước khi nhúng để file nhẹ
                    thumb_path = str(Path(screenshot).with_suffix(".thumb.png"))
                    with PILImage.open(screenshot) as im:
                        im.thumbnail((200, 130))
                        im.save(thumb_path)

                    xl_img = XLImage(thumb_path)
                    xl_img.width  = 200
                    xl_img.height = 130
                    col_letter = ws.cell(r_idx, self._SCREENSHOT_COL).column_letter
                    ws.add_image(xl_img, f"{col_letter}{r_idx}")
                    ws.row_dimensions[r_idx].height = self._ROW_HEIGHT_PX
                except Exception as e:
                    ws.cell(r_idx, self._SCREENSHOT_COL, f"[img error: {e}]")

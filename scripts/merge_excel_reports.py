"""
merge_excel_reports.py
──────────────────────
Gộp tất cả {module}_report.xlsx trong reports/excel/ thành all_report.xlsx.
Sheet Summary tổng hợp + 1 sheet riêng cho từng module.

Usage:
    python scripts/merge_excel_reports.py
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

EXCEL_DIR = Path("reports/excel")
OUTPUT    = EXCEL_DIR / "all_report.xlsx"

_MODULES = ["signalgenerator", "power", "amplifier", "attenuator", "spectrumanalyzer", "vna"]

# ── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _bold(size: int = 11) -> Font:
    return Font(bold=True, size=size)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _thin() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

_GREEN  = "FF92D050"
_RED    = "FFFF0000"
_YELLOW = "FFFFC000"
_BLUE   = "FF4472C4"
_GREY   = "FFD9D9D9"
_PURPLE = "FFB8CCE4"
_WHITE  = "FFFFFFFF"

# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    wb_out = Workbook()

    # ── Collect data từ từng module ──────────────────────────────────────────
    all_rows: list[tuple] = []   # (module, row_values_list)
    module_sheets: dict[str, list] = {}

    for module in _MODULES:
        src = EXCEL_DIR / f"{module}_report.xlsx"
        if not src.exists():
            print(f"  [SKIP] {src.name} — không tồn tại")
            continue

        wb_src = load_workbook(src, data_only=True)
        # Lấy sheet "Test Results" hoặc sheet đầu tiên không phải Summary
        sheet_name = "Test Results" if "Test Results" in wb_src.sheetnames else wb_src.sheetnames[-1]
        ws_src = wb_src[sheet_name]

        rows = []
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(row)
        module_sheets[module] = rows
        all_rows.extend(rows)
        print(f"  [OK]   {src.name} — {len(rows)} rows")

    if not module_sheets:
        print("Không có module report nào — bỏ qua.")
        return

    # ── Sheet Summary ────────────────────────────────────────────────────────
    ws_sum = wb_out.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 12

    def _sum_row(r, label, value, val_fill=None):
        lc = ws_sum.cell(r, 1, label)
        vc = ws_sum.cell(r, 2, value)
        lc.font = _bold(); lc.fill = _fill(_GREY)
        lc.border = _thin(); lc.alignment = _center()
        vc.border = _thin(); vc.alignment = _center()
        if val_fill:
            vc.fill = _fill(val_fill)

    def _count(rows, outcome):
        return sum(1 for r in rows if r[6] and str(r[6]).upper() == outcome.upper())

    total   = len(all_rows)
    passed  = _count(all_rows, "PASSED")
    failed  = _count(all_rows, "FAILED")
    skipped = _count(all_rows, "SKIPPED")
    manual  = _count(all_rows, "MANUAL")
    errors  = _count(all_rows, "ERROR")
    auto_total = total - manual

    _sum_row(1, "Report generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    _sum_row(2, "Modules",   len(module_sheets))
    _sum_row(3, "Total",     total)
    _sum_row(4, "PASSED",    passed,  _GREEN)
    _sum_row(5, "FAILED",    failed,  _RED)
    _sum_row(6, "SKIPPED",   skipped, _YELLOW)
    _sum_row(7, "MANUAL",    manual,  _PURPLE)
    _sum_row(8, "ERROR",     errors,  _RED)
    _sum_row(9, "Pass rate", f"{(passed/auto_total*100):.1f}%" if auto_total else "N/A")

    # Per-module summary
    ws_sum.cell(11, 1, "Module").font = _bold()
    ws_sum.cell(11, 2, "Total").font  = _bold()
    ws_sum.cell(11, 3, "Pass").font   = _bold()
    ws_sum.cell(11, 4, "Fail").font   = _bold()
    ws_sum.column_dimensions["C"].width = 10
    ws_sum.column_dimensions["D"].width = 10

    for ri, (mod, rows) in enumerate(module_sheets.items(), start=12):
        ws_sum.cell(ri, 1, mod)
        ws_sum.cell(ri, 2, len(rows))
        ws_sum.cell(ri, 3, _count(rows, "PASSED")).fill = _fill(_GREEN)
        ws_sum.cell(ri, 4, _count(rows, "FAILED")).fill = _fill(_RED) if _count(rows, "FAILED") else _fill(_WHITE)

    # ── Sheet per module ─────────────────────────────────────────────────────
    HEADERS = ["Test ID","Brief","Level","Type","Execution","HW Depend",
               "Result","Duration (s)","Error / Notes","Screenshot"]
    WIDTHS  = [20, 40, 12, 14, 16, 11, 10, 13, 50, 30]
    OUTCOME_FILL = {
        "PASSED": _fill(_GREEN), "FAILED": _fill(_RED),
        "SKIPPED": _fill(_YELLOW), "MANUAL": _fill(_PURPLE),
    }

    for module, rows in module_sheets.items():
        ws = wb_out.create_sheet(module)
        for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
            c = ws.cell(1, ci, h)
            c.font = Font(bold=True, color="FFFFFFFF", size=11)
            c.fill = _fill(_BLUE); c.alignment = _center(); c.border = _thin()
            ws.column_dimensions[c.column_letter].width = w
        ws.row_dimensions[1].height = 24

        for ri, row in enumerate(rows, start=2):
            outcome = str(row[6]).upper() if row[6] else ""
            row_fill = OUTCOME_FILL.get(outcome, _fill(_WHITE))
            for ci, val in enumerate(row[:10], start=1):
                c = ws.cell(ri, ci, val)
                c.border = _thin()
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if ci == 7:
                    c.fill = row_fill; c.font = _bold(); c.alignment = _center()

    wb_out.save(OUTPUT)
    print(f"\nAll report: {OUTPUT}  ({total} tests, {len(module_sheets)} modules)")


if __name__ == "__main__":
    main()

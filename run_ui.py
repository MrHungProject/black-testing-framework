"""
Black Testing Framework — Desktop UI
Chạy: python run_ui.py
"""

import glob
import os
import re
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import ttk

_BASE_DIR = os.path.dirname(
    sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
)
TESTS_DIR   = os.path.join(_BASE_DIR, "tests")
REPORTS_DIR = os.path.join(_BASE_DIR, "reports")

if getattr(sys, "frozen", False):
    import shutil
    PYTHON = shutil.which("python") or shutil.which("python3") or "python"
else:
    PYTHON = sys.executable

_TYPE_COLOR = {
    "fat":        "#fab387",
    "functional": "#89dceb",
    "mixed":      "#cba6f7",
    "unknown":    "#585b70",
}
_TYPE_LABEL = {
    "fat":        "FAT",
    "functional": "FUNC",
    "mixed":      "MIX",
    "unknown":    "?",
}
_OUTCOME_TAG = {
    "passed":  "pass",
    "failed":  "fail",
    "error":   "error",
    "skipped": "skip",
    "manual":  "manual",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def get_file_test_type(file_path: str) -> str:
    try:
        with open(file_path, encoding="utf-8", errors="ignore") as f:
            content = f.read()
        has_fat  = "@test_type: fat"        in content
        has_func = "@test_type: functional" in content
        if has_fat and has_func:
            return "mixed"
        if has_fat:
            return "fat"
        if has_func:
            return "functional"
    except Exception:
        pass
    return "unknown"


def discover_tests(root: str) -> dict[str, list[tuple[str, str]]]:
    result = {}
    for folder in sorted(os.listdir(root)):
        folder_path = os.path.join(root, folder)
        if not os.path.isdir(folder_path) or folder.startswith("_"):
            continue
        files = sorted(
            f for f in os.listdir(folder_path)
            if f.startswith("test_") and f.endswith(".py")
        )
        if files:
            result[folder] = [
                (f, get_file_test_type(os.path.join(folder_path, f)))
                for f in files
            ]
    return result


def discover_test_cases(file_path: str) -> list[tuple[str, str, str]]:
    """Parse test file → [(class_name, func_name, brief), ...]"""
    results = []
    try:
        with open(file_path, encoding="utf-8", errors="ignore") as f:
            content = f.read()

        current_class = ""
        for line in content.splitlines():
            cm = re.match(r'^class\s+(\w+)', line)
            if cm:
                current_class = cm.group(1)
            fm = re.match(r'^\s+def\s+(test_\w+)\s*\(', line)
            if fm and current_class:
                results.append((current_class, fm.group(1), ""))

        brief_map: dict[str, str] = {}
        for m in re.finditer(r'def\s+(test_\w+).*?@brief:\s*([^\n@]+)', content, re.DOTALL):
            brief_map[m.group(1)] = m.group(2).strip()

        results = [(cls, fn, brief_map.get(fn, "")) for cls, fn, _ in results]
    except Exception:
        pass
    return results


def load_excel_report(path: str) -> tuple[list[str], list[list], str]:
    """Đọc Excel report — tìm sheet có cột 'Test ID' và 'Result'.
    Trả về (headers, data_rows, error_msg). error_msg rỗng nếu thành công."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            lower   = [h.lower() for h in headers]
            if "result" in lower and "test id" in lower:
                data = [
                    [str(c) if c is not None else "" for c in row]
                    for row in all_rows[1:]
                ]
                wb.close()
                return headers, data, ""
        wb.close()
        return [], [], "Không tìm thấy sheet hợp lệ trong file."
    except ImportError:
        return [], [], "Thiếu thư viện openpyxl. Chạy: pip install openpyxl"
    except Exception as e:
        return [], [], f"Lỗi đọc file: {e}"


# ── Main App ───────────────────────────────────────────────────────────────────

class TestRunnerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Black Testing Framework")
        self.geometry("1280x760")
        self.minsize(900, 520)
        self.configure(bg="#1e1e2e")

        self._proc: subprocess.Popen | None = None
        self._running = False
        self._check_vars: dict[str, tk.BooleanVar] = {}
        self._tc_vars:    dict[str, tk.BooleanVar] = {}   # "module/file::Class::func" → BoolVar
        self._expanded_files: set[str] = set()            # file keys currently expanded
        self._filter_var = tk.StringVar(value="all")

        self._build_ui()
        self._load_tests()
        self._refresh_report()

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        title_frame = tk.Frame(self, bg="#181825", pady=8)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame, text="  Black Testing Framework",
            font=("Segoe UI", 14, "bold"), fg="#cdd6f4", bg="#181825"
        ).pack(side="left")

        main = tk.Frame(self, bg="#1e1e2e")
        main.pack(fill="both", expand=True)

        # ── Left sidebar ───────────────────────────────────────────────────────
        sidebar = tk.Frame(main, bg="#181825", width=310)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        filter_frame = tk.Frame(sidebar, bg="#181825", pady=6)
        filter_frame.pack(fill="x", padx=10)
        tk.Label(
            filter_frame, text="Loại test:", font=("Segoe UI", 9, "bold"),
            fg="#a6adc8", bg="#181825"
        ).pack(side="left")
        for value, label, color in [
            ("all",        "Tất cả",     "#cdd6f4"),
            ("fat",        "FAT",        "#fab387"),
            ("functional", "Functional", "#89dceb"),
        ]:
            tk.Radiobutton(
                filter_frame, text=label,
                variable=self._filter_var, value=value,
                bg="#181825", fg=color,
                activebackground="#181825", activeforeground=color,
                selectcolor="#313244",
                font=("Segoe UI", 9, "bold"),
                command=self._on_filter_change,
            ).pack(side="left", padx=(6, 0))

        tk.Frame(sidebar, bg="#313244", height=1).pack(fill="x", padx=8)
        tk.Label(
            sidebar, text="Test Modules", font=("Segoe UI", 10, "bold"),
            fg="#89b4fa", bg="#181825", pady=6
        ).pack(anchor="w", padx=12)

        self._canvas = tk.Canvas(sidebar, bg="#181825", highlightthickness=0)
        scrollbar = tk.Scrollbar(sidebar, orient="vertical", command=self._canvas.yview)
        self._sidebar_frame = tk.Frame(self._canvas, bg="#181825")
        self._sidebar_frame.bind(
            "<Configure>",
            lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all"))
        )
        self._canvas.create_window((0, 0), window=self._sidebar_frame, anchor="nw")
        self._canvas.configure(yscrollcommand=scrollbar.set)
        self._canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            self._canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self._canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ── Right area ─────────────────────────────────────────────────────────
        right_frame = tk.Frame(main, bg="#1e1e2e")
        right_frame.pack(side="left", fill="both", expand=True, padx=(1, 0))

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",          background="#1e1e2e", borderwidth=0)
        style.configure("TNotebook.Tab",
            background="#181825", foreground="#a6adc8",
            font=("Segoe UI", 10, "bold"),
            padding=[14, 5], borderwidth=0,
        )
        style.map("TNotebook.Tab",
            background=[("selected", "#1e1e2e")],
            foreground=[("selected", "#cdd6f4")],
        )

        self._notebook = ttk.Notebook(right_frame)
        self._notebook.pack(fill="both", expand=True, pady=(4, 0))

        # Tab 1: Console
        console_tab = tk.Frame(self._notebook, bg="#1e1e2e")
        self._notebook.add(console_tab, text="  Console  ")

        self._console = tk.Text(
            console_tab,
            bg="#11111b", fg="#cdd6f4",
            font=("Consolas", 10),
            wrap="word", state="disabled",
            relief="flat", bd=0,
        )
        console_scroll = tk.Scrollbar(console_tab, command=self._console.yview)
        self._console.configure(yscrollcommand=console_scroll.set)
        console_scroll.pack(side="right", fill="y")
        self._console.pack(fill="both", expand=True, padx=(8, 0), pady=4)

        self._console.tag_config("pass",   foreground="#a6e3a1")
        self._console.tag_config("fail",   foreground="#f38ba8")
        self._console.tag_config("error",  foreground="#fab387")
        self._console.tag_config("warn",   foreground="#f9e2af")
        self._console.tag_config("info",   foreground="#89dceb")
        self._console.tag_config("dim",    foreground="#585b70")
        self._console.tag_config("header", foreground="#cba6f7", font=("Consolas", 10, "bold"))

        # Tab 2: Report
        report_tab = tk.Frame(self._notebook, bg="#1e1e2e")
        self._notebook.add(report_tab, text="  Report  ")
        self._build_report_tab(report_tab)

        # Bottom toolbar
        toolbar = tk.Frame(self, bg="#181825", pady=6)
        toolbar.pack(fill="x", side="bottom")

        self._btn_run = tk.Button(
            toolbar, text="▶  Run Selected",
            font=("Segoe UI", 10, "bold"),
            bg="#89b4fa", fg="#1e1e2e",
            relief="flat", padx=16, pady=6,
            cursor="hand2", command=self._run_tests,
        )
        self._btn_run.pack(side="left", padx=(12, 6))

        self._btn_stop = tk.Button(
            toolbar, text="■  Stop",
            font=("Segoe UI", 10),
            bg="#f38ba8", fg="#1e1e2e",
            relief="flat", padx=12, pady=6,
            cursor="hand2", state="disabled",
            command=self._stop_tests,
        )
        self._btn_stop.pack(side="left", padx=4)

        tk.Button(
            toolbar, text="Clear",
            font=("Segoe UI", 10),
            bg="#313244", fg="#cdd6f4",
            relief="flat", padx=12, pady=6,
            cursor="hand2", command=self._clear_console,
        ).pack(side="left", padx=4)

        self._status_var = tk.StringVar(value="Ready")
        tk.Label(
            toolbar, textvariable=self._status_var,
            font=("Segoe UI", 9), fg="#a6adc8", bg="#181825"
        ).pack(side="right", padx=16)

    # ── Report tab ─────────────────────────────────────────────────────────────

    def _build_report_tab(self, parent: tk.Frame):
        top = tk.Frame(parent, bg="#181825", pady=8)
        top.pack(fill="x", padx=8, pady=(8, 0))

        tk.Label(
            top, text="Report", font=("Segoe UI", 11, "bold"),
            fg="#89b4fa", bg="#181825"
        ).pack(side="left", padx=(4, 12))

        tk.Label(top, text="File:", font=("Segoe UI", 9),
                 fg="#a6adc8", bg="#181825").pack(side="left")
        self._report_file_var = tk.StringVar()
        self._report_combo = ttk.Combobox(
            top, textvariable=self._report_file_var,
            state="readonly", width=28, font=("Consolas", 9),
        )
        self._report_combo.pack(side="left", padx=(4, 10))
        self._report_combo.bind("<<ComboboxSelected>>", lambda *_: self._load_report_table())

        for label, cmd, color in [
            ("🔄 Refresh",     self._refresh_report,    "#313244"),
            ("📂 Excel",       self._open_excel,        "#313244"),
            ("🌐 HTML",        self._open_html,         "#313244"),
            ("📋 Log",         self._open_log,          "#313244"),
            ("📸 Screenshots", self._open_screenshots,  "#313244"),
        ]:
            tk.Button(
                top, text=label, font=("Segoe UI", 9),
                bg=color, fg="#cdd6f4",
                relief="flat", padx=10, pady=4,
                cursor="hand2", command=cmd,
            ).pack(side="left", padx=3)

        summary_row = tk.Frame(parent, bg="#1e1e2e")
        summary_row.pack(fill="x", padx=8, pady=(6, 0))
        self._lbl_summary = tk.Label(
            summary_row, text="Chưa có report. Chạy test để tạo.",
            font=("Segoe UI", 9), fg="#a6adc8", bg="#1e1e2e", anchor="w"
        )
        self._lbl_summary.pack(fill="x", padx=4)

        tree_frame = tk.Frame(parent, bg="#1e1e2e")
        tree_frame.pack(fill="both", expand=True, padx=8, pady=(6, 0))

        style = ttk.Style()
        style.configure("Report.Treeview",
            background="#11111b", foreground="#cdd6f4",
            fieldbackground="#11111b", rowheight=24,
            font=("Consolas", 9),
        )
        style.configure("Report.Treeview.Heading",
            background="#181825", foreground="#89b4fa",
            font=("Segoe UI", 9, "bold"), relief="flat",
        )
        style.map("Report.Treeview",
            background=[("selected", "#313244")],
            foreground=[("selected", "#ffffff")],
        )

        cols = ("test_id", "brief", "type", "outcome", "duration", "error")
        self._tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            selectmode="browse", style="Report.Treeview"
        )
        self._tree.heading("test_id",  text="Test ID")
        self._tree.heading("brief",    text="Brief")
        self._tree.heading("type",     text="Type")
        self._tree.heading("outcome",  text="Kết quả")
        self._tree.heading("duration", text="Thời gian")
        self._tree.heading("error",    text="Lỗi")

        self._tree.column("test_id",  width=250, minwidth=150, stretch=False)
        self._tree.column("brief",    width=300, minwidth=100)
        self._tree.column("type",     width=75,  minwidth=50,  stretch=False, anchor="center")
        self._tree.column("outcome",  width=80,  minwidth=60,  stretch=False, anchor="center")
        self._tree.column("duration", width=70,  minwidth=50,  stretch=False, anchor="center")
        self._tree.column("error",    width=280, minwidth=80)

        self._tree.tag_configure("pass",   foreground="#a6e3a1")
        self._tree.tag_configure("fail",   foreground="#f38ba8")
        self._tree.tag_configure("error",  foreground="#fab387")
        self._tree.tag_configure("skip",   foreground="#6c7086")
        self._tree.tag_configure("manual", foreground="#cba6f7")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)
        self._tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        detail_frame = tk.Frame(parent, bg="#181825")
        detail_frame.pack(fill="x", padx=8, pady=(4, 6))
        tk.Label(
            detail_frame, text="Chi tiết lỗi:",
            font=("Segoe UI", 8), fg="#6c7086", bg="#181825"
        ).pack(anchor="w", padx=4)
        self._detail_text = tk.Text(
            detail_frame,
            bg="#181825", fg="#fab387",
            font=("Consolas", 9),
            height=4, wrap="word",
            state="disabled", relief="flat",
        )
        self._detail_text.pack(fill="x", padx=4, pady=(0, 4))

    # ── Report logic ───────────────────────────────────────────────────────────

    def _get_excel_files(self) -> list[str]:
        excel_dir = os.path.join(REPORTS_DIR, "excel")
        if not os.path.isdir(excel_dir):
            return []
        return sorted(glob.glob(os.path.join(excel_dir, "*.xlsx")), reverse=True)

    def _refresh_report(self):
        files = self._get_excel_files()
        names = [os.path.basename(f) for f in files]
        self._report_combo["values"] = names
        if names:
            self._report_file_var.set(names[0])
            self._load_report_table()
        else:
            self._lbl_summary.config(text="Chưa có report. Chạy test để tạo.")

    def _load_report_table(self):
        fname = self._report_file_var.get()
        if not fname:
            return
        path = os.path.join(REPORTS_DIR, "excel", fname)
        headers, rows, err_msg = load_excel_report(path)

        for item in self._tree.get_children():
            self._tree.delete(item)
        self._detail_text.config(state="normal")
        self._detail_text.delete("1.0", "end")
        self._detail_text.config(state="disabled")

        if not headers:
            self._lbl_summary.config(text=err_msg or "Không đọc được report.")
            return

        if not rows:
            self._lbl_summary.config(text="Report rỗng — chưa có test nào được chạy.")
            return

        lower = [h.lower() for h in headers]

        def col(name: str) -> int:
            try:
                return lower.index(name.lower())
            except ValueError:
                return -1

        ci_id      = col("test id")
        ci_brief   = col("brief")
        ci_type    = col("type")
        ci_outcome = col("result")
        ci_dur     = col("duration (s)")
        ci_err     = col("error / notes")

        def get(row, c):
            return row[c] if 0 <= c < len(row) else ""

        counts = {"passed": 0, "failed": 0, "error": 0, "skipped": 0, "manual": 0}

        for row in rows:
            outcome = get(row, ci_outcome).lower()
            tag     = _OUTCOME_TAG.get(outcome, "skip")
            if outcome in counts:
                counts[outcome] += 1
            err_short = get(row, ci_err).replace("\n", " ")[:150]
            self._tree.insert(
                "", "end",
                values=(
                    get(row, ci_id),
                    get(row, ci_brief),
                    get(row, ci_type),
                    outcome.upper(),
                    get(row, ci_dur),
                    err_short,
                ),
                tags=(tag,),
            )

        total = sum(counts.values())
        p = counts["passed"]
        f = counts["failed"] + counts["error"]
        s = counts["skipped"]
        m = counts["manual"]
        self._lbl_summary.config(
            text=(
                f"  {fname}  |  Total: {total}    "
                f"✓ {p} passed    ✗ {f} failed    "
                f"↷ {s} skipped    ✦ {m} manual"
            )
        )

    def _on_tree_select(self, _event):
        sel = self._tree.selection()
        if not sel:
            return
        values = self._tree.item(sel[0], "values")
        err = values[5] if len(values) > 5 else ""
        self._detail_text.config(state="normal")
        self._detail_text.delete("1.0", "end")
        if err:
            self._detail_text.insert("1.0", err)
        self._detail_text.config(state="disabled")

    def _open_excel(self):
        fname = self._report_file_var.get()
        if fname:
            path = os.path.join(REPORTS_DIR, "excel", fname)
            if os.path.exists(path):
                os.startfile(path)
                return
        excel_dir = os.path.join(REPORTS_DIR, "excel")
        if os.path.isdir(excel_dir):
            os.startfile(excel_dir)

    def _open_html(self):
        html = os.path.join(REPORTS_DIR, "html", "report.html")
        if os.path.exists(html):
            os.startfile(html)
        else:
            self._log("HTML report chưa có.\n", "warn")

    def _open_log(self):
        log = os.path.join(REPORTS_DIR, "logs", "framework.log")
        if os.path.exists(log):
            os.startfile(log)
        else:
            self._log("Log file chưa có.\n", "warn")

    def _open_screenshots(self):
        ss_dir = os.path.join(REPORTS_DIR, "screenshots")
        if os.path.isdir(ss_dir):
            os.startfile(ss_dir)
        else:
            self._log("Screenshots chưa có.\n", "warn")

    # ── Load / filter tests ────────────────────────────────────────────────────

    def _on_filter_change(self):
        self._load_tests()

    def _load_tests(self):
        selected_filter = self._filter_var.get()
        all_tests = discover_tests(TESTS_DIR)

        for widget in self._sidebar_frame.winfo_children():
            widget.destroy()
        self._check_vars.clear()
        self._tc_vars.clear()

        any_visible = False
        for module, file_entries in all_tests.items():
            visible_files = (
                file_entries if selected_filter == "all"
                else [(f, t) for f, t in file_entries if t == selected_filter or t == "mixed"]
            )
            if not visible_files:
                continue

            any_visible = True
            mod_var = tk.BooleanVar(value=False)
            self._check_vars[module] = mod_var

            # Module header row
            mod_frame = tk.Frame(self._sidebar_frame, bg="#181825")
            mod_frame.pack(fill="x", padx=8, pady=(8, 0))
            tk.Checkbutton(
                mod_frame, text=f"  {module}",
                variable=mod_var, bg="#181825",
                fg="#cdd6f4", activebackground="#181825",
                activeforeground="#cdd6f4", selectcolor="#313244",
                font=("Segoe UI", 10, "bold"),
                command=lambda m=module, v=mod_var: self._toggle_module(m, v),
            ).pack(side="left")

            # File rows
            for fname, ftype in visible_files:
                file_key = f"{module}/{fname}"
                fvar = tk.BooleanVar(value=False)
                self._check_vars[file_key] = fvar

                # Outer frame wraps file row + tc container
                outer = tk.Frame(self._sidebar_frame, bg="#181825")
                outer.pack(fill="x", padx=0, pady=0)

                row = tk.Frame(outer, bg="#181825")
                row.pack(fill="x")

                tc_container = tk.Frame(outer, bg="#181825")

                is_exp = [file_key in self._expanded_files]

                expand_lbl = tk.Label(
                    row,
                    text="▼" if is_exp[0] else "▶",
                    font=("Consolas", 9), fg="#6c7086", bg="#181825",
                    cursor="hand2", padx=4,
                )
                expand_lbl.pack(side="left")

                def _make_toggle(fk, m, f, lbl, cont, fv, exp_flag):
                    def _toggle(*_):
                        if exp_flag[0]:
                            exp_flag[0] = False
                            self._expanded_files.discard(fk)
                            lbl.config(text="▶")
                            cont.pack_forget()
                            for k in list(self._tc_vars.keys()):
                                if k.startswith(fk + "::"):
                                    del self._tc_vars[k]
                        else:
                            exp_flag[0] = True
                            self._expanded_files.add(fk)
                            lbl.config(text="▼")
                            self._populate_tc_list(fk, m, f, cont, fv)
                            cont.pack(fill="x")
                        self._canvas.configure(
                            scrollregion=self._canvas.bbox("all")
                        )
                    return _toggle

                toggle_fn = _make_toggle(
                    file_key, module, fname, expand_lbl, tc_container, fvar, is_exp
                )
                expand_lbl.bind("<Button-1>", toggle_fn)

                tk.Checkbutton(
                    row,
                    text=f"  {fname.replace('.py', '')}",
                    variable=fvar, bg="#181825",
                    fg="#a6adc8", activebackground="#181825",
                    activeforeground="#cdd6f4", selectcolor="#313244",
                    font=("Consolas", 9), anchor="w",
                    command=lambda fk=file_key, fv=fvar: self._on_file_check_change(fk, fv.get()),
                ).pack(side="left", fill="x", expand=True)

                if selected_filter == "all":
                    tk.Label(
                        row,
                        text=_TYPE_LABEL.get(ftype, "?"),
                        font=("Consolas", 7, "bold"),
                        fg="#1e1e2e", bg=_TYPE_COLOR.get(ftype, "#585b70"),
                        padx=3, pady=0,
                    ).pack(side="right", padx=(0, 4))

                # If already expanded, populate immediately
                if is_exp[0]:
                    self._populate_tc_list(file_key, module, fname, tc_container, fvar)
                    tc_container.pack(fill="x")

        if not any_visible:
            tk.Label(
                self._sidebar_frame,
                text=f"  Không có test {selected_filter.upper()}",
                font=("Segoe UI", 9), fg="#585b70", bg="#181825",
            ).pack(anchor="w", padx=16, pady=12)

    def _populate_tc_list(
        self,
        file_key: str,
        module: str,
        fname: str,
        container: tk.Frame,
        file_var: tk.BooleanVar,
    ):
        for widget in container.winfo_children():
            widget.destroy()

        file_path = os.path.join(TESTS_DIR, module, fname)
        tcs = discover_test_cases(file_path)

        if not tcs:
            tk.Label(
                container, text="    (không tìm thấy test case)",
                fg="#585b70", bg="#181825", font=("Segoe UI", 8),
            ).pack(anchor="w", padx=28, pady=2)
            return

        file_checked = file_var.get()
        for class_name, func_name, brief in tcs:
            tc_key = f"{file_key}::{class_name}::{func_name}"
            existing = self._tc_vars.get(tc_key)
            tc_var = existing if existing is not None else tk.BooleanVar(value=file_checked)
            self._tc_vars[tc_key] = tc_var

            tc_row = tk.Frame(container, bg="#181825")
            tc_row.pack(fill="x", pady=0)

            tk.Checkbutton(
                tc_row,
                variable=tc_var, bg="#181825",
                selectcolor="#313244", activebackground="#181825",
            ).pack(side="left", padx=(28, 0))

            num = func_name.split("_")[-1]
            display = f"{num}  {brief[:48]}" if brief else func_name
            tk.Label(
                tc_row, text=display,
                font=("Consolas", 8), fg="#6c7086", bg="#181825",
                anchor="w",
            ).pack(side="left", fill="x", expand=True, padx=(2, 6))

    def _on_file_check_change(self, file_key: str, checked: bool):
        """Sync TC vars when file checkbox changes."""
        for tc_key, tc_var in self._tc_vars.items():
            if tc_key.startswith(file_key + "::"):
                tc_var.set(checked)

    def _toggle_module(self, module: str, mod_var: tk.BooleanVar):
        checked = mod_var.get()
        for key, var in self._check_vars.items():
            if key.startswith(f"{module}/"):
                var.set(checked)
        for tc_key, tc_var in self._tc_vars.items():
            if tc_key.startswith(f"{module}/"):
                tc_var.set(checked)

    # ── Run / Stop ─────────────────────────────────────────────────────────────

    def _collect_selected_paths(self) -> list[str]:
        """
        Build the pytest target list.
        - Expanded files: use individual checked TC node IDs
        - Non-expanded files: use whole file path if file checkbox is checked
        - Module checkbox selects all files in module
        """
        selected: list[str] = []
        files_handled: set[str] = set()

        # 1. Individual TC selections (from expanded files)
        for tc_key, tc_var in self._tc_vars.items():
            if not tc_var.get():
                continue
            slash = tc_key.index("/")
            dcolon = tc_key.index("::")
            module = tc_key[:slash]
            fname  = tc_key[slash + 1:dcolon]
            file_key = f"{module}/{fname}"
            class_func = tc_key[dcolon + 2:]        # "ClassName::test_func"
            node_id = os.path.join(TESTS_DIR, module, fname) + "::" + class_func
            selected.append(node_id)
            files_handled.add(file_key)

        # 2. Whole-file selections (non-expanded or file not in tc_vars yet)
        for key, var in self._check_vars.items():
            if not var.get():
                continue
            parts = key.split("/")
            if len(parts) == 1:
                continue  # module-level checkbox — handled via file-level
            module, fname = parts[0], parts[1]
            file_key = f"{module}/{fname}"
            if file_key in files_handled:
                continue
            if file_key in self._expanded_files:
                continue  # expanded → TC level controls it
            selected.append(os.path.join(TESTS_DIR, module, fname))

        return selected

    def _run_tests(self):
        paths = self._collect_selected_paths()
        if not paths:
            self._log("Chưa chọn test nào!\n", "warn")
            return

        cmd = [PYTHON, "-m", "pytest", "-v", "--tb=short", "--no-header"]
        flt = self._filter_var.get()
        if flt in ("fat", "functional"):
            cmd += ["-m", flt]
        cmd += paths

        self._notebook.select(0)

        self._log("=" * 60 + "\n", "header")
        filter_label = {"all": "ALL", "fat": "FAT only", "functional": "Functional only"}.get(flt, flt)
        tc_count = sum(1 for p in paths if "::" in os.path.basename(p) or "::" in p)
        file_count = len(paths) - tc_count
        if tc_count:
            self._log(f"  Chạy {tc_count} test case cụ thể\n", "info")
        if file_count:
            self._log(f"  Chạy {file_count} file\n", "info")
        self._log(f"  Filter: {filter_label}\n", "info")
        self._log(f"CMD: {' '.join(cmd)}\n", "dim")
        self._log("=" * 60 + "\n", "header")

        self._running = True
        self._btn_run.config(state="disabled")
        self._btn_stop.config(state="normal")
        self._status_var.set(f"Running [{filter_label}]...")

        threading.Thread(target=self._stream_output, args=(cmd,), daemon=True).start()

    def _stream_output(self, cmd: list[str]):
        try:
            self._proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace",
                cwd=_BASE_DIR,
            )
            for line in self._proc.stdout:
                self._classify_and_log(line)
            self._proc.wait()
            rc = self._proc.returncode
            self._log("\n" + "=" * 60 + "\n", "header")
            if rc == 0:
                self._log("  ALL PASSED ✓\n", "pass")
            else:
                self._log(f"  SOME FAILED ✗  (exit code {rc})\n", "fail")
            self._log("=" * 60 + "\n", "header")
        except Exception as e:
            self._log(f"[ERROR] {e}\n", "error")
        finally:
            self._running = False
            self.after(0, self._on_done)

    def _classify_and_log(self, line: str):
        l = line.lower()
        if " passed" in l and "failed" not in l:
            tag = "pass"
        elif "failed" in l or " error" in l:
            tag = "fail"
        elif "warning" in l:
            tag = "warn"
        elif line.startswith("PASSED"):
            tag = "pass"
        elif line.startswith("FAILED"):
            tag = "fail"
        elif line.startswith("ERROR"):
            tag = "error"
        elif line.startswith("=") or line.startswith("-"):
            tag = "header"
        else:
            tag = None
        self._log(line, tag)

    def _on_done(self):
        self._btn_run.config(state="normal")
        self._btn_stop.config(state="disabled")
        self._status_var.set("Done")
        self._refresh_report()
        self._notebook.select(1)

    def _stop_tests(self):
        if self._proc:
            self._proc.terminate()
        self._status_var.set("Stopped")

    def _log(self, text: str, tag: str | None = None):
        def _write():
            self._console.config(state="normal")
            if tag:
                self._console.insert("end", text, tag)
            else:
                self._console.insert("end", text)
            self._console.see("end")
            self._console.config(state="disabled")
        self.after(0, _write)

    def _clear_console(self):
        self._console.config(state="normal")
        self._console.delete("1.0", "end")
        self._console.config(state="disabled")
        self._status_var.set("Ready")


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = TestRunnerApp()
    app.mainloop()
